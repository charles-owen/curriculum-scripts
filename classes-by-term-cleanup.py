"""
Cleanup of the student information system classes by term
spreadsheet, This is for processing the output of the query
MSU_RPT_CLASSES_BY_TERM_NOINST

Usage:
    classes-by-term-cleanup <report>

Options:
    <report>    The report generated .xlsx file
"""

from docopt import docopt
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles.fills import PatternFill
from openpyxl.workbook.views import BookView
import re
import warnings
import time
import os

class Settings:
    def __init__(self, args):
        self.file = args['<report>']


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    args = docopt(__doc__)
    # print(args)

    settings = Settings(args)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(settings.file)

    sheet = wb['sheet1']

    # Delete the first row
    sheet.delete_rows(1, 1)

    # Get the headings
    headings = []
    for cell in sheet[1]:
        headings.append(cell.value)

    # print(headings)

    #
    # Delete unwanted columns
    #
    to_delete = ['Term', 'Course ID', 'Career', 'Session', 'Offer Nbr', 'Component', 'Min Units', 'Class Nbr',
                 'Pat Nbr', 'Class Type', 'Class Assoc', 'Location', 'Unit Acad Org', 'Unit Acad Org Descr',
                 'MAU Acad Org', 'MAU Acad Org Descr']

    for delete in to_delete:
        i = headings.index(delete)
        if i is not None:
            del headings[i]
            sheet.delete_cols(i+1, 1)


    #
    # Set column widths
    #
    widths = [['Class Title', 29], ['Subject', 6.67], ['Section', 6.67], ['Mode', 5.17],
              ['Meeting Start Time', 8.83], ['Meeting End Time', 8.83],
              ['Cap Enrl', 6.0], ['Tot Enrl', 6.0], ['Wait Tot', 6.0], ['Room Cap', 9.33],
              ['Open Seats', 9.83]]
    for width in widths:
        i = headings.index(width[0])
        if i is not None:
            sheet.column_dimensions[get_column_letter(i+1)].width = width[1]
        else:
            print(f'Column {width[0]} not found')

    colors = [['^\\s*102', 'fffa80'], ['^\\s*220', 'e0e0e0'], ['^\\s*231', 'ffcc80'], ['^\\s*232', 'ff80e6'],
             ['^\\s*260', '808dff'], ['^\\s*3', '80ff80'], ['^\\s*490', 'ffffff'],
              ['^\\s*498', 'ff8080'], ['^\\s*4', 'ff8093']]
    catalog = headings.index('Catalog')
    for row in range(2, sheet.max_row+1):
        course = sheet[row][catalog]
        # print(f'"{course.value}"')
        gotIt = False
        for color in colors:
            if re.search(color[0], course.value) is not None:
                # print(f'Set {course.value} to {color[1]}')
                gotIt = True

                for rows1 in sheet.iter_rows(min_row=row, max_row=row, min_col=1, max_col=len(headings)):
                    for cell in rows1:
                        cell.fill = PatternFill(start_color=color[1], end_color=color[1], fill_type="solid")
                break

            if gotIt:
                break

    # Set the sheet zoom level
    sheet.sheet_view.zoomScale = 135

    sheet.freeze_panes = sheet['A2']

    # Make it open a more reasonable size
    view = [BookView(windowWidth=36000, windowHeight=36000)]
    wb.views = view

    # ['Term', 'Term Desc', , 'Subject', 'Catalog', 'Section',
    # , 'Max Units', 'Facil ID', 'Day Codes', 'Meeting Start Time', 'Meeting End Time',
    # , 'Enrl Stat', 'Class Stat', 'Mode', 'Cap Enrl', 'Tot Enrl',
    # 'Wait Tot', 'Room Cap', 'Open Seats',
    # 'Class Title']

    dir = os.path.dirname(settings.file)
    file = os.path.basename(settings.file)
    output_file = os.path.join(dir, time.strftime("%Y-%m-%d-") + file)
    wb.save(output_file)


