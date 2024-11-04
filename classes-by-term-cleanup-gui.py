"""
Cleanup of the student information system classes by term
spreadsheet, This is for processing the output of the query
MSU_RPT_CLASSES_BY_TERM_NOINST
"""

from guizero import App, PushButton, Box
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles.fills import PatternFill
from openpyxl.workbook.views import BookView
import re
import warnings
import time
import os
from copy import copy

def sheet_sort_rows(ws, row_start, row_end=0, cols=None, sorter=None, reverse=False):
    """ Sorts given rows of the sheet
        From: https://stackoverflow.com/questions/44767554/sorting-with-openpyxl
        row_start   First row to be sorted
        row_end     Last row to be sorted (default last row)
        cols        Columns to be considered in sort
        sorter      Function that accepts a tuple of values and
                    returns a sortable key
        reverse     Reverse the sort order
    """

    bottom = ws.max_row
    if row_end == 0:
        row_end = ws.max_row
    right = get_column_letter(ws.max_column)
    if cols is None:
        cols = range(1, ws.max_column + 1)

    array = {}
    for row in range(row_start, row_end + 1):
        key = ''
        for col in cols:
            key += ws.cell(row, col).value
        array[key] = array.get(key, set()).union({row})

    order = sorted(array, key=sorter, reverse=reverse)

    ws.move_range(f"A{row_start}:{right}{row_end}", bottom)
    dest = row_start
    for src_key in order:
        for row in array[src_key]:
            src = row + bottom
            dist = dest - src
            ws.move_range(f"A{src}:{right}{src}", dist)
            dest += 1

class ClassesByTerm(App):
    def __init__(self):
        super(ClassesByTerm, self).__init__(title="Classes By Term", width=300, height=200)
        box = Box(self, align="left", width="fill")
        button = PushButton(box, text="Open output of\nMSU_RPT_CLASSES_BY_TERM_NOINST\nquery...", command=self.on_open, args=[self], align="top", width=25, height=4)
        exit = PushButton(box, text="Exit", command=self.on_exit, args=[self], align="top", width=25, height=2)
        pass

    def on_exit(self, value):
        app.destroy()

    def on_open(self, value):
        file = app.select_file(filetypes=[["All files", "*.*"], ["Excel files", "*.xlsx"]])
        if file is None:
            return

        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(file)

        sheet = wb['sheet1']

        # Delete the first row
        sheet.delete_rows(1, 1)

        # Get the headings
        headings = []
        for cell in sheet[1]:
            headings.append(cell.value)

        # print(headings)

        #
        # Delete unwanted columns
        #
        to_delete = ['Term', 'Course ID', 'Career', 'Session', 'Offer Nbr', 'Component', 'Min Units', 'Class Nbr',
                     'Pat Nbr', 'Class Type', 'Class Assoc', 'Location', 'Unit Acad Org', 'Unit Acad Org Descr',
                     'MAU Acad Org', 'MAU Acad Org Descr']

        for delete in to_delete:
            i = headings.index(delete)
            if i is not None:
                del headings[i]
                sheet.delete_cols(i + 1, 1)

        #
        # Set column widths
        #
        widths = [['Class Title', 29], ['Subject', 6.67], ['Section', 6.67], ['Mode', 5.17],
                  ['Meeting Start Time', 8.83], ['Meeting End Time', 8.83],
                  ['Cap Enrl', 6.0], ['Tot Enrl', 6.0], ['Wait Tot', 6.0], ['Room Cap', 9.33],
                  ['Open Seats', 9.83]]
        for width in widths:
            i = headings.index(width[0])
            if i is not None:
                sheet.column_dimensions[get_column_letter(i + 1)].width = width[1]
            else:
                print(f'Column {width[0]} not found')

        colors = [['^\\s*102', 'fffa80', 'faf8c5'], ['^\\s*220', 'e0e0e0', 'e0e0e0'],
                  ['^\\s*231', 'ffcc80', 'ffcc80'],
                  ['^\\s*232', 'ff80e6', 'ffcff5'],
                  ['^\\s*260', '808dff', '808dff'],
                  ['^\\s*3', '80ff80', 'cfffcf'],
                  ['^\\s*490', 'ffffff', 'ffffff'],
                  ['^\\s*498', 'ff8080', 'ff8080'],
                  ['^\\s*4', '80dbff', 'd9f4ff']]
        cancelled_color = '888888'
        catalog = headings.index('Catalog')
        enrl_status = headings.index('Enrl Stat')
        class_status = headings.index('Class Stat')
        for row in range(2, sheet.max_row + 1):
            course = sheet[row][catalog]
            color_to_use = 'ffffff'
            font_color = None

            for color in colors:
                if re.search(color[0], course.value) is not None:
                    color_to_use = color[1]
                    cancelled_color = color[2]
                    break

            # Test for cancelled sections
            if class_status is not None:
                status = sheet[row][class_status].value
                if status.startswith('Cancelled') or status.startswith('Stop Further'):
                    color_to_use = cancelled_color
                    font_color = '888888'
                elif enrl_status is not None:
                    e_status = sheet[row][enrl_status].value
                    if e_status.startswith('Closed'):
                        font_color = 'ff0000'

            for rows1 in sheet.iter_rows(min_row=row, max_row=row, min_col=1, max_col=len(headings)):
                for cell in rows1:
                    cell.fill = PatternFill(start_color=color_to_use, end_color=color_to_use, fill_type="solid")

                    if font_color is not None:
                        new_font = copy(cell.font)
                        new_font.color = font_color
                        cell.font = new_font


        # Sort the rows
        sheet_sort_rows(sheet, 2, sheet.max_row, [3, 4, 2])

        # Set the sheet zoom level
        sheet.sheet_view.zoomScale = 135

        sheet.freeze_panes = sheet['A2']

        # Make it open a more reasonable size
        view = [BookView(windowWidth=36000, windowHeight=36000)]
        wb.views = view

        # ['Term', 'Term Desc', , 'Subject', 'Catalog', 'Section',
        # , 'Max Units', 'Facil ID', 'Day Codes', 'Meeting Start Time', 'Meeting End Time',
        # , 'Enrl Stat', 'Class Stat', 'Mode', 'Cap Enrl', 'Tot Enrl',
        # 'Wait Tot', 'Room Cap', 'Open Seats',
        # 'Class Title']

        dir = os.path.dirname(file)
        filename = time.strftime("%Y-%m-%d-") + os.path.basename(file)
        output_file = os.path.join(dir, filename)
        wb.save(output_file)

        app.info("Success", "File successfully saved as " + filename)





#
# Program entry point
#
if __name__ == '__main__':
    app = ClassesByTerm()
    app.display()

